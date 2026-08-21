#!/usr/bin/env python3
"""Map every color in palettes.json to the nearest Ohuhu Honolulu marker color.

Reads the marker database xlsx and the original palettes, then writes:
  - markers.json         : the parsed marker list
  - marker_palettes.json : palettes with each color replaced by its nearest marker

Usage:
  python build_marker_palettes.py <marker-database.xlsx>

Nearest match uses CIEDE2000 distance in CIELAB space. Within a single
palette, each marker is used at most once: if the nearest marker is already
taken, the next-nearest is used, so a palette never shows duplicate markers.
"""

import json
import math
import sys
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).parent


def parse_markers(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    markers = []
    # Data starts at row 4: No., swatch, old code, old name, new code, new name, hex, RGB, HSL
    for row in ws.iter_rows(min_row=4, values_only=True):
        no, _swatch, old_code, old_name, code, name, hex_code, _rgb, _hsl = row
        if no is None or hex_code is None:
            continue
        hex_code = str(hex_code).strip().lower()
        if not (hex_code.startswith("#") and len(hex_code) == 7):
            raise ValueError(f"Row {no}: bad hex {hex_code!r}")
        markers.append(
            {
                "code": str(code).strip(),
                "oldCode": str(old_code).strip(),
                "name": str(name).strip(),
                "oldName": str(old_name).strip(),
                "hex": hex_code,
            }
        )
    return markers


def hex_to_rgb(hex_code):
    return tuple(int(hex_code[i : i + 2], 16) for i in (1, 3, 5))


def rgb_to_lab(rgb):
    # sRGB -> linear -> XYZ (D65) -> CIELAB
    def to_linear(c):
        c /= 255.0
        return c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4

    r, g, b = (to_linear(c) for c in rgb)
    x = (0.4124564 * r + 0.3575761 * g + 0.1804375 * b) / 0.95047
    y = 0.2126729 * r + 0.7151522 * g + 0.0721750 * b
    z = (0.0193339 * r + 0.1191920 * g + 0.9503041 * b) / 1.08883

    def f(t):
        return t ** (1 / 3) if t > 0.008856 else 7.787 * t + 16 / 116

    fx, fy, fz = f(x), f(y), f(z)
    return (116 * fy - 16, 500 * (fx - fy), 200 * (fy - fz))


def ciede2000(lab1, lab2):
    L1, a1, b1 = lab1
    L2, a2, b2 = lab2
    C1 = math.hypot(a1, b1)
    C2 = math.hypot(a2, b2)
    C_avg = (C1 + C2) / 2
    G = 0.5 * (1 - math.sqrt(C_avg**7 / (C_avg**7 + 25**7)))
    a1p, a2p = (1 + G) * a1, (1 + G) * a2
    C1p, C2p = math.hypot(a1p, b1), math.hypot(a2p, b2)
    h1p = math.degrees(math.atan2(b1, a1p)) % 360 if (a1p or b1) else 0
    h2p = math.degrees(math.atan2(b2, a2p)) % 360 if (a2p or b2) else 0

    dLp = L2 - L1
    dCp = C2p - C1p
    if C1p * C2p == 0:
        dhp = 0
    else:
        dh = h2p - h1p
        if dh > 180:
            dh -= 360
        elif dh < -180:
            dh += 360
        dhp = dh
    dHp = 2 * math.sqrt(C1p * C2p) * math.sin(math.radians(dhp) / 2)

    Lp_avg = (L1 + L2) / 2
    Cp_avg = (C1p + C2p) / 2
    if C1p * C2p == 0:
        hp_avg = h1p + h2p
    else:
        diff = abs(h1p - h2p)
        s = h1p + h2p
        if diff <= 180:
            hp_avg = s / 2
        elif s < 360:
            hp_avg = (s + 360) / 2
        else:
            hp_avg = (s - 360) / 2

    T = (
        1
        - 0.17 * math.cos(math.radians(hp_avg - 30))
        + 0.24 * math.cos(math.radians(2 * hp_avg))
        + 0.32 * math.cos(math.radians(3 * hp_avg + 6))
        - 0.20 * math.cos(math.radians(4 * hp_avg - 63))
    )
    d_theta = 30 * math.exp(-(((hp_avg - 275) / 25) ** 2))
    Rc = 2 * math.sqrt(Cp_avg**7 / (Cp_avg**7 + 25**7))
    Sl = 1 + 0.015 * (Lp_avg - 50) ** 2 / math.sqrt(20 + (Lp_avg - 50) ** 2)
    Sc = 1 + 0.045 * Cp_avg
    Sh = 1 + 0.015 * Cp_avg * T
    Rt = -math.sin(math.radians(2 * d_theta)) * Rc

    return math.sqrt(
        (dLp / Sl) ** 2
        + (dCp / Sc) ** 2
        + (dHp / Sh) ** 2
        + Rt * (dCp / Sc) * (dHp / Sh)
    )


def main():
    if len(sys.argv) != 2:
        sys.exit("Usage: build_marker_palettes.py <marker-database.xlsx>")

    markers = parse_markers(sys.argv[1])
    print(f"Parsed {len(markers)} markers")

    marker_labs = [rgb_to_lab(hex_to_rgb(m["hex"])) for m in markers]
    palettes = json.loads((PROJECT_DIR / "palettes.json").read_text())

    match_cache = {}  # original hex -> markers sorted by distance

    def ranked_matches(hex_code):
        if hex_code not in match_cache:
            lab = rgb_to_lab(hex_to_rgb(hex_code))
            distances = [
                (ciede2000(lab, mlab), i) for i, mlab in enumerate(marker_labs)
            ]
            distances.sort()
            match_cache[hex_code] = distances
        return match_cache[hex_code]

    out = []
    for palette in palettes:
        used = set()
        entries = []
        for hex_code in palette:
            hex_code = hex_code.lower()
            for dist, i in ranked_matches(hex_code):
                if i not in used:
                    used.add(i)
                    m = markers[i]
                    entries.append(
                        {
                            "hex": m["hex"],
                            "code": m["code"],
                            "name": m["name"],
                            "oldCode": m["oldCode"],
                            "oldName": m["oldName"],
                            "originalHex": hex_code,
                            "deltaE": round(dist, 2),
                        }
                    )
                    break
        out.append(entries)

    (PROJECT_DIR / "markers.json").write_text(json.dumps(markers, indent=1))
    (PROJECT_DIR / "marker_palettes.json").write_text(json.dumps(out))
    unique_used = {e["code"] for p in out for e in p}
    delta_es = [e["deltaE"] for p in out for e in p]
    print(f"Wrote {len(out)} palettes -> marker_palettes.json")
    print(f"Markers used across all palettes: {len(unique_used)}/{len(markers)}")
    print(
        f"deltaE avg {sum(delta_es)/len(delta_es):.2f}, "
        f"max {max(delta_es):.2f}"
    )


if __name__ == "__main__":
    main()
